from datetime import datetime
from io import BytesIO
import logging
import math
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties

_logger = logging.getLogger(__name__)


def _should_save_pdf_snapshot() -> bool:
    return str(os.environ.get("EXCEL2PDF_SAVE_LOCAL_PDF", "1")).strip() == "1"


def _save_pdf_snapshot_if_needed(pdf_bytes: bytes, utils_dir: Path):
    if not _should_save_pdf_snapshot():
        return
    snapshot_path = utils_dir / f"excel2pdf_output_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.pdf"
    snapshot_path.write_bytes(pdf_bytes)


def fix_excel_layout_for_pdf(wb):
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        merged_ranges = ws.merged_cells.ranges
        for row in ws.iter_rows():
            max_lines = 1
            is_header = row[0].row == 1
            for cell in row:
                if cell.value:
                    text = str(cell.value)
                    if cell.alignment:
                        cell.alignment = cell.alignment.copy(wrapText=True, vertical="center")
                    else:
                        cell.alignment = Alignment(wrapText=True, vertical="center")
                    if cell.font and cell.font.size and cell.font.size > 10:
                        cell.font = cell.font.copy(size=10)
                    actual_width = 0
                    for merged_range in merged_ranges:
                        if cell.coordinate in merged_range:
                            for col in range(merged_range.min_col, merged_range.max_col + 1):
                                col_letter = openpyxl.utils.get_column_letter(col)
                                actual_width += ws.column_dimensions[col_letter].width or 8.43
                            break
                    else:
                        actual_width = ws.column_dimensions[cell.column_letter].width or 8.43
                    manual_lines = text.count("\n") + 1
                    chars_per_line = max(actual_width * 1.0, 5)
                    auto_lines = math.ceil(len(text) / chars_per_line)
                    current_lines = max(manual_lines, auto_lines)
                    max_lines = max(max_lines, current_lines)
            calculated_height = max_lines * 18 + 16
            if is_header:
                calculated_height += 10
            ws.row_dimensions[row[0].row].height = max(calculated_height, 22)


def _generated_pdf_bytes(tmp_path: Path, input_path: Path):
    generated_pdf = input_path.with_suffix(".pdf")
    if generated_pdf.exists():
        return generated_pdf.read_bytes(), generated_pdf

    alt_generated_pdf = tmp_path / f"{input_path.name}.pdf"
    if alt_generated_pdf.exists():
        return alt_generated_pdf.read_bytes(), alt_generated_pdf

    candidate_pdfs = sorted(
        [p for p in tmp_path.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if candidate_pdfs:
        return candidate_pdfs[0].read_bytes(), candidate_pdfs[0]

    return None, generated_pdf


def _generated_pdf_from_output(output_text: str, tmp_path: Path):
    for matched_pdf in re.findall(r"->\s*(.+?\.pdf)\s+using filter", output_text, flags=re.IGNORECASE):
        candidate = Path(matched_pdf.strip()).resolve()
        if candidate.exists() and tmp_path.resolve() in candidate.parents:
            return candidate.read_bytes()
    return None


def _should_try_xvfb(detail: str, returncode: int | None = None) -> bool:
    lowered = (detail or "").lower()
    return (
        returncode == 134
        or "signal 11" in lowered
        or "fatal exception" in lowered
        or "can't open display" in lowered
        or "x11 error" in lowered
    )


def convert_excel_bytes_to_pdf_bytes(excel_bytes: bytes, excel_suffix: str = ".xlsx") -> bytes:
    normalized_suffix = excel_suffix.lower()
    if normalized_suffix not in {".xls", ".xlsx", ".xlsm"}:
        raise ValueError("excel_suffix 必须是 .xls/.xlsx/.xlsm")

    utils_dir = Path(__file__).resolve().parent
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        input_path = tmp_path / f"input{normalized_suffix}"
        libreoffice_home = tmp_path / "libreoffice-home"
        libreoffice_home.mkdir(parents=True, exist_ok=True)
        libreoffice_profile = (tmp_path / "libreoffice-profile").resolve()

        try:
            wb = openpyxl.load_workbook(BytesIO(excel_bytes))
            fix_excel_layout_for_pdf(wb)
            optimized_bytes = BytesIO()
            wb.save(optimized_bytes)
            input_path.write_bytes(optimized_bytes.getvalue())
        except Exception:
            input_path.write_bytes(excel_bytes)

        env = os.environ.copy()
        env["LANG"] = "C.UTF-8"
        env["LC_ALL"] = "C.UTF-8"
        env["HOME"] = str(libreoffice_home)
        env["USER"] = "ragflow"

        env.pop("DISPLAY", None)

        env["SAL_USE_VCLPLUGIN"] = "svp"
        env["SAL_DISABLE_GLX"] = "1"

        env["LD_LIBRARY_PATH"] = ":".join([
            "/usr/lib/libreoffice/program",
            "/usr/lib",
            "/usr/lib/x86_64-linux-gnu",
            env.get("LD_LIBRARY_PATH", "")
        ]).strip(":")
        env["UNO_PATH"] = "/usr/lib/libreoffice/program"
        env["PATH"] = "/usr/lib/libreoffice/program:" + env["PATH"]

        office_bin = "/usr/bin/libreoffice" if Path("/usr/bin/libreoffice").exists() else (shutil.which("libreoffice") or shutil.which("soffice"))
        if not office_bin:
            _logger.error(
                "[excel2pdf] LibreOffice 不可用：未找到 libreoffice/soffice（已检查 /usr/bin/libreoffice 与 PATH）。"
                "Excel 转 PDF 无法执行，请在运行解析任务的环境中安装 LibreOffice（Docker 镜像需包含对应包）。"
            )
            raise RuntimeError("未找到 LibreOffice/soffice 可执行文件，请在容器内安装 libreoffice。")

        command = [
            office_bin,
            "--headless",
            "--invisible",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--nofirststartwizard",
            f"-env:UserInstallation=file://{libreoffice_profile.as_posix()}",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmp_path),
            str(input_path),
        ]

        def _run_convert(run_env, run_command=None):
            return subprocess.run(run_command or command, check=True, capture_output=True, text=True, env=run_env)

        def _run_convert_with_xvfb(detail):
            xvfb_run = shutil.which("xvfb-run")
            if not xvfb_run:
                raise RuntimeError(
                    "Excel 转 PDF 失败，LibreOffice 在容器内异常退出，且系统中没有可用的 xvfb-run。"
                    "请在镜像里安装 xvfb、libreoffice、libreoffice-calc、fonts-noto-cjk、locales-all 后重试。"
                    f"原始输出: {detail}。"
                )

            xvfb_env = env.copy()
            xvfb_env.pop("DISPLAY", None)
            xvfb_env["SAL_USE_VCLPLUGIN"] = "gen"
            xvfb_env["SAL_DISABLE_GLX"] = "1"
            xvfb_command = [
                xvfb_run,
                "-a",
                "-s",
                "-screen 0 1024x768x24",
                *command,
            ]
            return _run_convert(xvfb_env, xvfb_command)

        try:
            result = _run_convert(env)
        except subprocess.CalledProcessError as exc:
            stderr = (exc.stderr or "").strip()
            stdout = (exc.stdout or "").strip()
            detail = stderr or stdout or "无额外输出"
            if _should_try_xvfb(detail, exc.returncode):
                try:
                    result = _run_convert_with_xvfb(detail)
                except subprocess.CalledProcessError as retry_exc:
                    retry_stderr = (retry_exc.stderr or "").strip()
                    retry_stdout = (retry_exc.stdout or "").strip()
                    retry_detail = retry_stderr or retry_stdout or "无额外输出"
                    raise RuntimeError(
                        f"Excel 转 PDF 失败，LibreOffice 直接运行异常退出(exit={exc.returncode})，"
                        f"并且在 xvfb-run 下重试仍失败(exit={retry_exc.returncode})。"
                        f"命令: {' '.join(command)}。"
                        f"首轮输出: {detail}。"
                        f"xvfb 重试输出: {retry_detail}。"
                    ) from retry_exc
            else:
                raise RuntimeError(
                    f"Excel 转 PDF 失败(exit={exc.returncode})。"
                    f"命令: {' '.join(command)}。"
                    f"输出: {detail}。"
                    "常见原因：容器缺少 libreoffice 运行时依赖，或者当前用户没有可写的配置目录。"
                ) from exc

        pdf_bytes, generated_pdf = _generated_pdf_bytes(tmp_path, input_path)
        if pdf_bytes is not None:
            _save_pdf_snapshot_if_needed(pdf_bytes, utils_dir)
            return pdf_bytes

        output_text = f"{(result.stdout or '').strip()}\n{(result.stderr or '').strip()}".strip()
        pdf_bytes = _generated_pdf_from_output(output_text, tmp_path)
        if pdf_bytes is not None:
            _save_pdf_snapshot_if_needed(pdf_bytes, utils_dir)
            return pdf_bytes

        detail = output_text or "无额外输出"
        if _should_try_xvfb(detail):
            try:
                retry_result = _run_convert_with_xvfb(detail)
            except subprocess.CalledProcessError as retry_exc:
                retry_stderr = (retry_exc.stderr or "").strip()
                retry_stdout = (retry_exc.stdout or "").strip()
                retry_detail = retry_stderr or retry_stdout or "无额外输出"
                raise RuntimeError(
                    f"Excel 转 PDF 失败(xvfb-run 重试仍失败，exit={retry_exc.returncode})。"
                    f"命令: {' '.join(command)}。"
                    f"首轮输出: {detail}。"
                    f"重试输出: {retry_detail}。"
                ) from retry_exc

            pdf_bytes, generated_pdf = _generated_pdf_bytes(tmp_path, input_path)
            if pdf_bytes is not None:
                _save_pdf_snapshot_if_needed(pdf_bytes, utils_dir)
                return pdf_bytes

            retry_output_text = f"{(retry_result.stdout or '').strip()}\n{(retry_result.stderr or '').strip()}".strip()
            pdf_bytes = _generated_pdf_from_output(retry_output_text, tmp_path)
            if pdf_bytes is not None:
                _save_pdf_snapshot_if_needed(pdf_bytes, utils_dir)
                return pdf_bytes
            detail = retry_output_text or detail

        raise RuntimeError(
            f"转换失败，未生成输出文件: {generated_pdf}。"
            f"命令: {' '.join(command)}。"
            f"输出: {detail}。"
        )


def convert_excel_to_pdf(
    input_excel: str,
    output_pdf: str | None = None,
) -> str:
    input_path = Path(input_excel).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if input_path.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
        raise ValueError("输入文件必须是 .xls/.xlsx/.xlsm")
    if output_pdf is None:
        output_path = input_path.with_suffix(".pdf")
    else:
        output_path = Path(output_pdf).expanduser().resolve()
    if output_path.suffix.lower() != ".pdf":
        raise ValueError("输出文件必须是 .pdf")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    excel_bytes = input_path.read_bytes()
    pdf_bytes = convert_excel_bytes_to_pdf_bytes(
        excel_bytes=excel_bytes,
        excel_suffix=input_path.suffix,
    )
    output_path.write_bytes(pdf_bytes)
    return str(output_path)


if __name__ == "__main__":
    current_dir = Path(__file__).resolve().parent
    test_excel = current_dir / "测试.xlsx"
    test_pdf = current_dir / "test.pdf"
    result_pdf = convert_excel_to_pdf(
        input_excel=str(test_excel),
        output_pdf=str(test_pdf),
    )
    print(f"转换完成: {result_pdf}")
